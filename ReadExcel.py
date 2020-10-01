import openpyxl
path = "D:\Pic\empRecord.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active
#sheet = workbook.get_sheet_by_name("sheet1")

r = sheet.max_row      #6
c = sheet.max_column   #4

print(r)
print(c)

for r in range (1,r+1):
    for c in range(1,c+1):
        print(sheet.cell(row=r,column=c).value, end="     ")
    print()